import streamlit as st
from openpyxl import load_workbook
import time


class ApplicationUI:
    """
    Classe com os objetos utilziados para
    criar o frontned da aplcição de unir as bases
    """

    def __init__(self) -> None:
        self.init_session()
        self.start_layout()
    
        if st.session_state['base_file_uploaded'] == False:
            self.first_file = self.file_loader("Base")

            if self.first_file is not None:
                st.session_state['base_file_uploaded'] = self.check_valid_base_file(self.first_file)
                
                if st.session_state['base_file_uploaded'] == True:
                    st.session_state['first_file'] = self.first_file
                    st.success('Boa! A base foi importada, agora suba a planilha de novos trajetos.')
                    time.sleep(3)
                    st.rerun()

        if st.session_state['base_file_uploaded'] and st.session_state['append_file_uploaded'] == False:
            self.second_file = self.file_loader("Novos Trajetos")
            
            if self.second_file is not None:
                st.session_state['append_file_uploaded'] = self.check_valid_base_file(self.second_file)

                if st.session_state['append_file_uploaded'] == True:
                    st.session_state['second_file'] = self.second_file
                    st.success('Boa! A planilha de trajetos foi importada, aguarde o processamento.')
                    time.sleep(3)
                    st.rerun()
        
        if st.session_state['base_file_uploaded'] == True and st.session_state['append_file_uploaded'] == True:
            self.start_button()

    def check_valid_base_file(self, file):
        if load_workbook(file).active.max_column > 5:
            return True
        return False

    def start_layout(self) -> None:
        st.title('Atualizar Base')
        st.write('Instruções: Insira a Base atual e o arquivo mais recente de trajetos e clique em atualizar base. Espere o processamento (~2 minutos) e quando liberado, clique em baixar arquivos. Caso algum erro ocorra, as instruções serão evidenciadas.')

    def file_loader(self, label) -> None:
        st.subheader(f'{label}')
        st.write(f'Faça o Upload da {label.lower()}')
        return st.file_uploader(label=label, type=['.xlsx', '.xls'], accept_multiple_files=False)

    def start_button(self) -> None:
        st.subheader('Boa!')
        st.write('Você já subiu ambas as planilhas necessárias. Clique em processar e aguarde o download do arquivo.')
        st.button('Processar')
        st.download_button(label='Download', file_name='Nova Base.xlsx', data=st.session_state['first_file'])

    def user_success_message(self) -> None:
        pass

    def user_erros_message(self) -> None:
        pass

    def init_session(self) -> None:
        if 'base_file_uploaded' not in st.session_state:
            st.session_state['base_file_uploaded'] = False
        if 'append_file_uploaded' not in st.session_state:
            st.session_state['append_file_uploaded'] = False
        if 'first_file' not in st.session_state:
            st.session_state['first_file'] = None
        if 'second_file' not in st.session_state:
            st.session_state['second_file'] = None


teste = ApplicationUI()
